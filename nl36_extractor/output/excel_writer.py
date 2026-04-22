"""
Excel Writer for NL-36 Distribution of Business via Intermediaries.

Output:
  - Master_Data sheet: one row per (company, quarter, year, channel)
  - Per-company verification sheet: channels as rows, 8 period-metric cols
  - Validation_Summary and Validation_Detail sheets
"""

import logging
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import MASTER_COLUMNS, NUMBER_FORMAT, company_key_to_pascal
from config.channel_registry import CHANNEL_DISPLAY_NAMES, CHANNEL_ORDER
from extractor.models import NL36Extract

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_META_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# Data columns (H onward) — match MASTER_COLUMNS order
_DATA_FIELDS = [
    "cy_qtr_policies", "cy_qtr_premium",
    "cy_ytd_policies", "cy_ytd_premium",
    "py_qtr_policies", "py_qtr_premium",
    "py_ytd_policies", "py_ytd_premium",
]

_DATA_HEADERS = [
    "CY Qtr\nPolicies", "CY Qtr\nPremium (₹L)",
    "CY YTD\nPolicies", "CY YTD\nPremium (₹L)",
    "PY Qtr\nPolicies", "PY Qtr\nPremium (₹L)",
    "PY YTD\nPolicies", "PY YTD\nPremium (₹L)",
]


def _sheet_name_for(ext: NL36Extract) -> str:
    """Build verification sheet name (max 31 chars)."""
    name = f"{company_key_to_pascal(ext.company_key)}_{ext.quarter}_{ext.year}"
    return name[:31]


def _write_master_data(ws, extractions: List[NL36Extract]) -> None:
    """Write Master_Data sheet — one row per (company, quarter, year, channel)."""
    for ci, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    next_row = 2
    for ext in extractions:
        channel_map = {ch.channel_key: ch for ch in ext.channels}
        for ch_key in CHANNEL_ORDER:
            ch = channel_map.get(ch_key)
            if ch is None:
                continue
            row_data = [
                ext.company_name,
                ext.company_key,
                ext.quarter,
                ext.year,
                ext.source_file,
                ch.channel_key,
                CHANNEL_DISPLAY_NAMES.get(ch.channel_key, ch.channel_key),
            ]
            for field in _DATA_FIELDS:
                row_data.append(getattr(ch, field))

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=ci, value=val)
                if ci >= 8 and val is not None:
                    cell.number_format = NUMBER_FORMAT
            next_row += 1

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)


def _write_company_sheet(ws, ext: NL36Extract) -> None:
    """Write per-company verification sheet: channels as rows, 8 data cols."""
    # Title
    title_cell = ws.cell(row=1, column=1, value=ext.company_name)
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    subtitle = f"NL-36  |  Quarter: {ext.quarter}  |  Year: {ext.year}  |  Source: {ext.source_file}"
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    # Header row
    ws.cell(row=4, column=1, value="Distribution Channel").font = Font(bold=True)
    ws.cell(row=4, column=1).fill = _HEADER_FILL
    ws.cell(row=4, column=1).font = _HEADER_FONT
    ws.cell(row=4, column=1).alignment = _CENTER_ALIGN

    for ci, hdr in enumerate(_DATA_HEADERS, 2):
        cell = ws.cell(row=4, column=ci, value=hdr)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN

    ws.freeze_panes = "A5"

    # Data rows
    channel_map = {ch.channel_key: ch for ch in ext.channels}
    current_row = 5
    for ch_key in CHANNEL_ORDER:
        ch = channel_map.get(ch_key)
        if ch is None:
            continue

        label = CHANNEL_DISPLAY_NAMES.get(ch_key, ch_key)
        is_total = ch_key in ("total_channel", "grand_total")

        label_cell = ws.cell(row=current_row, column=1, value=label)
        if is_total:
            label_cell.font = Font(bold=True)
            label_cell.fill = _META_FILL

        for ci, field in enumerate(_DATA_FIELDS, 2):
            val = getattr(ch, field)
            cell = ws.cell(row=current_row, column=ci, value=val)
            if val is not None:
                cell.number_format = NUMBER_FORMAT
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = _META_FILL

        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 36
    for col in "BCDEFGHI":
        ws.column_dimensions[col].width = 18


def save_workbook(
    extractions: List[NL36Extract],
    output_path: str,
    stats: Optional[dict] = None,
) -> None:
    """Write all extractions to the master Excel workbook."""
    # Load existing or create fresh
    output_file = Path(output_path)
    if output_file.exists():
        wb = load_workbook(output_path)
        # Remove sheets being replaced
        new_files = {e.source_file for e in extractions}
        if "Master_Data" in wb.sheetnames:
            del wb["Master_Data"]
        for ext in extractions:
            sn = _sheet_name_for(ext)
            if sn in wb.sheetnames:
                del wb[sn]
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 1. Master_Data (always first)
    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions)

    # 2. Per-company verification sheets
    for ext in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(ext))
        _write_company_sheet(ws, ext)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path}")


def write_validation_summary_sheet(report_path: str, workbook_path: str, force_company: str = None) -> None:
    """
    Append a Validation_Summary sheet — one row per company with
    PASS / WARN / FAIL counts as columns (mirrors NL35 layout).
    """
    import pandas as pd

    if not Path(report_path).exists():
        return

    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status",
        aggfunc="size",
        fill_value=0,
    ).reset_index()

    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0

    summary["Files_Processed"] = 1
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    summary["Total_Checks"] = summary[["PASS", "SKIP", "WARN", "FAIL"]].sum(axis=1)
    cols = ["Company", "Quarter", "Year", "Files_Processed", "Total_Checks", "PASS", "SKIP", "WARN", "FAIL"]
    summary = summary[cols]
    if force_company:
        try:
            existing = pd.read_excel(workbook_path, sheet_name="Validation_Summary")
            companies_in_new = set(summary["Company"].unique())
            existing = existing[~existing["Company"].isin(companies_in_new)]
            summary = pd.concat([existing, summary], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(workbook_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, workbook_path: str, force_company: str = None) -> None:
    """
    Append a Validation_Detail sheet — FAILs and WARNs only, with
    renamed columns, sorted by Status, and red/yellow row highlights
    (mirrors NL35 layout).
    """
    import pandas as pd

    if not Path(report_path).exists():
        return

    df = pd.read_csv(report_path)
    cols_map = {
        "company":    "Company",
        "quarter":    "Quarter",
        "year":       "Year",
        "channel":    "Channel",
        "field":      "Field",
        "check_name": "Check_Name",
        "status":     "Status",
        "expected":   "Expected",
        "actual":     "Actual",
        "delta":      "Delta",
        "note":       "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values(by="Status").reset_index(drop=True)

    if force_company:
        try:
            run_companies = set(pd.read_csv(report_path)["company"].unique())
            existing_detail = pd.read_excel(workbook_path, sheet_name="Validation_Detail")
            if "Company" in existing_detail.columns:
                existing_detail = existing_detail[~existing_detail["Company"].isin(run_companies)]
            detail = pd.concat([existing_detail, detail], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(workbook_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        detail.to_excel(writer, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(workbook_path)
    ws = wb["Validation_Detail"]
    red_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col  = list(cols_map.values()).index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if status_val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(workbook_path)
    logger.info(f"Validation_Detail sheet written to {workbook_path}")
